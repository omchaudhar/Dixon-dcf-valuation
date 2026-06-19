"""
Dixon Technologies (India) Ltd.
3-Statement Financial Model + DCF Valuation  — v2.0
====================================================
Author  : Om Chaudhari | IIT Bombay B.Tech (Year 2)
Purpose : WiDS 5.0 Financial Modelling Bootcamp portfolio project
          Covers W1 (3-stmts + ratios), W2 (forecasting), W3 (WACC/DCF),
          W4 (peer comps), W5 (Python automation)
Currency: INR Crore (₹ Cr)
FY      : April 1 – March 31  |  NSE: DIXON  |  BSE: 541988

Data Sources:
  - Dixon Annual Reports FY22-FY25 (BSE/NSE filings)
    https://www.bseindia.com/bseplus/AnnualReport/541988/10124541988.pdf
  - Screener.in: https://www.screener.in/company/DIXON/consolidated/
  - Trendlyne:   https://trendlyne.com/equity/1624/DIXON/dixon-technologies-india-ltd/
  - Damodaran ERP India: https://pages.stern.nyu.edu/~adamodar/New_Home_Page/datafile/ctryprem.html
  - RBI 10-yr G-Sec yield (Rf): https://www.rbi.org.in/
  - ICICI Direct / Motilal Oswal research reports (broker estimates FY26-FY30)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
wb.remove(wb.active)

# ─── PALETTE & FORMATS ───────────────────────────────────────────────────────
DARK_BLUE    = "1F3864"
MED_BLUE     = "2E75B6"
LIGHT_BLUE   = "BDD7EE"
XLIGHT_BLUE  = "DEEAF1"
YELLOW       = "FFFF00"
LIGHT_YELLOW = "FFF2CC"
GREEN_FILL   = "E2EFDA"
GREY         = "F2F2F2"
ORANGE       = "F4B942"

INR   = '#,##0;(#,##0);"-"'
INR1  = '#,##0.0;(#,##0.0);"-"'
PCT   = '0.0%;(0.0%);"-"'
PCT2  = '0.00%;(0.00%);"-"'
MULT  = '0.0x'
PLAIN = '0.0'

def h_font(bold=True, color="FFFFFF", size=10):
    return Font(name='Calibri', bold=bold, color=color, size=size)

def fill(color):
    return PatternFill('solid', fgColor=color)

def al(h='center', v='center', wrap=False, indent=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)

def thin_border(bottom_only=False):
    s = Side(style='thin')
    if bottom_only:
        return Border(bottom=s)
    return Border(left=s, right=s, top=s, bottom=s)

def double_border():
    return Border(top=Side(style='thin'), bottom=Side(style='double'))

def set_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def inp(ws, row, col, val, fmt=INR, yellow=False):
    """Blue hardcoded input cell"""
    c = ws.cell(row=row, column=col, value=val)
    c.font  = Font(name='Calibri', bold=False, color="0070C0", size=10)
    c.number_format = fmt
    c.alignment = al('right')
    if yellow:
        c.fill = fill(LIGHT_YELLOW)
    return c

def frm(ws, row, col, formula, fmt=INR, color="000000"):
    """Formula cell — black by default, green if cross-sheet"""
    c = ws.cell(row=row, column=col, value=formula)
    c.font  = Font(name='Calibri', bold=False, color=color, size=10)
    c.number_format = fmt
    c.alignment = al('right')
    return c

def lbl(ws, row, col, text, bold=False, indent=1, bg=None, color="000000"):
    c = ws.cell(row=row, column=col, value=text)
    c.font  = Font(name='Calibri', bold=bold, color=color, size=10)
    c.alignment = al('left', indent=indent)
    if bg:
        c.fill = fill(bg)
    return c

def title_row(ws, row, text, bg=DARK_BLUE, fc="FFFFFF", size=12, cols=10):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font  = Font(name='Calibri', bold=True, color=fc, size=size)
    c.fill  = fill(bg)
    c.alignment = al('left', indent=1)
    ws.row_dimensions[row].height = 20
    return c

def sub_hdr(ws, row, text, bg=MED_BLUE, cols=10):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font  = Font(name='Calibri', bold=True, color="FFFFFF", size=10)
    c.fill  = fill(bg)
    c.alignment = al('left', indent=1)
    ws.row_dimensions[row].height = 16
    return c

def col_hdr(ws, row, labels, start_col=1, bg=DARK_BLUE):
    for i, txt in enumerate(labels):
        c = ws.cell(row=row, column=start_col+i, value=txt)
        c.font  = Font(name='Calibri', bold=True, color="FFFFFF", size=10)
        c.fill  = fill(bg)
        c.alignment = al('center')

def stripe(ws, row, cols=10, color=XLIGHT_BLUE):
    for c in range(1, int(cols)+1):
        ws.cell(row=row, column=c).fill = fill(color)

def totals_row(ws, row, cols=10):
    for c in range(1, int(cols)+1):
        ws.cell(row=row, column=c).border = double_border()

# Year column mapping  A=Labels, B=FY22A, C=FY23A, D=FY24A, E=FY25A, F=FY26E .. J=FY30E
YR_COLS  = ['B','C','D','E','F','G','H','I','J']
YR_LABELS= ['FY22A','FY23A','FY24A','FY25A','FY26E','FY27E','FY28E','FY29E','FY30E']
N_COLS   = 10   # A through J

# ─────────────────────────────────────────────────────────────────────────────
# TAB 1 — COVER
# ─────────────────────────────────────────────────────────────────────────────
cov = wb.create_sheet("Cover")
cov.sheet_view.showGridLines = False
set_widths(cov, {'A':3,'B':28,'C':22,'D':22,'E':22,'F':18,'G':18,'H':18,'I':18,'J':3})

cov.merge_cells('B2:I3')
t = cov['B2']
t.value = 'DIXON TECHNOLOGIES (INDIA) LTD.  |  NSE: DIXON  |  BSE: 541988'
t.font  = Font(name='Calibri', bold=True, color='FFFFFF', size=22)
t.fill  = fill(DARK_BLUE)
t.alignment = al('center')
cov.row_dimensions[2].height = 40
cov.row_dimensions[3].height = 40

cov.merge_cells('B4:I4')
s = cov['B4']
s.value = '3-Statement Financial Model & Discounted Cash Flow (DCF) Valuation'
s.font  = Font(name='Calibri', bold=True, color='FFFFFF', size=13)
s.fill  = fill(MED_BLUE)
s.alignment = al('center')
cov.row_dimensions[4].height = 24

meta = [
    ("Analyst",        "Om Chaudhari"),
    ("Institution",    "Indian Institute of Technology, Bombay — B.Tech (Year 2)"),
    ("Date",           "June 2026"),
    ("Currency",       "₹ Indian Rupee  |  Unit: Crore (₹ Cr)"),
    ("Fiscal Year",    "April 1 – March 31  (Indian FY)"),
    ("Exchange",       "NSE / BSE  |  ISIN: INE424L01029"),
    ("Scope",          "FY22A–FY25A (Actual)  +  FY26E–FY30E (Forecast)"),
    ("Model Version",  "v2.0  (8-tab structure, WACC merged into DCF)"),
]
for i, (k, v) in enumerate(meta, start=6):
    cov.cell(row=i, column=2, value=k).font   = Font(name='Calibri', bold=True, color=DARK_BLUE, size=10)
    cov.cell(row=i, column=3, value=v).font   = Font(name='Calibri', color="000000", size=10)
    cov.cell(row=i, column=2).alignment = al('left', indent=1)
    cov.cell(row=i, column=3).alignment = al('left', indent=1)
    if i % 2 == 0:
        for c in range(2, 10):
            cov.cell(row=i, column=c).fill = fill(XLIGHT_BLUE)

# Company Description
cov.row_dimensions[15].height = 6
title_row(cov, 16, '  COMPANY OVERVIEW', MED_BLUE, cols=8)
cov.merge_cells('B17:I22')
ov = cov['B17']
ov.value = (
    "Dixon Technologies (India) Ltd. is India's largest Electronics Manufacturing Services (EMS) company. "
    "Founded in 1993 and listed on NSE & BSE, Dixon manufactures across six verticals: Consumer Electronics "
    "(LED TVs, monitors), Home Appliances (washing machines), Lighting Products (LED lamps/luminaires), "
    "Mobile & EMS (smartphones under PLI scheme), Security Surveillance (CCTV/DVRs), and Wearables & IT Hardware "
    "(laptops, tablets). Dixon is a primary beneficiary of India's Production-Linked Incentive (PLI) schemes and "
    "the 'China+1' global supply-chain diversification theme. Key OEM partners include Samsung, Motorola, OnePlus, "
    "Xiaomi, Google, and HP. Management has guided for ₹1,00,000 Cr+ revenue by FY27."
)
ov.font = Font(name='Calibri', size=10)
ov.alignment = Alignment(wrap_text=True, vertical='top', indent=1)

# Key Financials
cov.row_dimensions[23].height = 6
title_row(cov, 24, '  KEY FINANCIAL HIGHLIGHTS  (₹ Crore)', MED_BLUE, cols=8)
col_hdr(cov, 25, ['Metric','FY22A','FY23A','FY24A','FY25A'], start_col=2)
kf = [
    ("Revenue (₹ Cr)",       "10,697","12,192","17,691","38,860"),
    ("YoY Revenue Growth",   "—",     "+14%",  "+45%",  "+120%"),
    ("EBITDA (₹ Cr)",        "430",   "510",   "708",   "1,943"),
    ("EBITDA Margin %",      "4.0%",  "4.2%",  "4.0%",  "5.0%"),
    ("PAT (₹ Cr)",           "190",   "253",   "365",   "1,215"),
    ("PAT Margin %",         "1.8%",  "2.1%",  "2.1%",  "3.1%"),
    ("EPS (₹)",              "31.7",  "42.2",  "60.9",  "202.8"),
    ("ROE",                  "17%",   "19%",   "21%",   "42%"),
]
for ri, rd in enumerate(kf, start=26):
    cov.cell(row=ri, column=2, value=rd[0]).font = Font(name='Calibri', bold=True, size=10)
    cov.cell(row=ri, column=2).alignment = al('left', indent=1)
    for ci, v in enumerate(rd[1:], start=3):
        c = cov.cell(row=ri, column=ci, value=v)
        c.font = Font(name='Calibri', size=10)
        c.alignment = al('center')
    if ri % 2 == 0:
        for c in range(2, 8):
            cov.cell(row=ri, column=c).fill = fill(XLIGHT_BLUE)

# Workbook Structure
cov.row_dimensions[35].height = 6
title_row(cov, 36, '  WORKBOOK STRUCTURE  (8 Tabs)', MED_BLUE, cols=8)
nav = [
    ("1. Cover",       "This page — project overview, key metrics, navigation guide"),
    ("2. ASSUMP",      "All model inputs & assumptions — edit ONLY this tab to change forecasts"),
    ("3. IS",          "Income Statement — P&L with Ratio Analysis (WiDS Week 1 & 2)"),
    ("4. BS",          "Balance Sheet — formula-linked to IS & CF"),
    ("5. CF",          "Cash Flow Statement — indirect method, formula-linked"),
    ("6. DCF",         "WACC derivation + FCFF model + EV bridge → Implied Share Price (Weeks 3-4)"),
    ("7. SENSITIVITY", "2-Way sensitivity table: Implied Price vs WACC & Terminal Growth Rate"),
    ("8. COMPS",       "Peer Comparables — Dixon vs Amber, Kaynes, Syrma (relative valuation)"),
]
for ri, (sh, desc) in enumerate(nav, start=37):
    cov.cell(row=ri, column=2, value=sh).font   = Font(name='Calibri', bold=True, color=MED_BLUE, size=10)
    cov.cell(row=ri, column=3, value=desc).font = Font(name='Calibri', size=10)
    cov.cell(row=ri, column=2).alignment = al('left', indent=1)
    cov.cell(row=ri, column=3).alignment = al('left', indent=1)
    if ri % 2 == 1:
        for c in range(2, 9):
            cov.cell(row=ri, column=c).fill = fill(XLIGHT_BLUE)

# Color Legend
cov.row_dimensions[46].height = 6
title_row(cov, 47, '  COLOR CODING CONVENTION', MED_BLUE, cols=8)
legend = [
    ("0070C0", "Blue text",        "Hardcoded inputs — source: Annual Reports / Damodaran / RBI"),
    ("000000", "Black text",       "Formula cells — calculated automatically, do NOT edit"),
    ("008000", "Green text",       "Cross-sheet links — pulling values from another tab"),
    (LIGHT_YELLOW[:-2], "Yellow fill", "Key assumption cell — verify before presenting"),
]
for ri, (clr, lbl_text, desc) in enumerate(legend, start=48):
    cov.cell(row=ri, column=2, value="  ■  " + lbl_text).font = Font(name='Calibri', bold=True, color=clr if len(clr)==6 else "FFC000", size=10)
    cov.cell(row=ri, column=3, value=desc).font = Font(name='Calibri', size=10)
    cov.cell(row=ri, column=2).alignment = al('left', indent=1)
    cov.cell(row=ri, column=3).alignment = al('left', indent=1)

# Sources
cov.row_dimensions[53].height = 6
src = cov.cell(row=54, column=2,
    value="Sources: Dixon Annual Reports FY22–FY25 (BSE/NSE) | Screener.in | Trendlyne | "
          "Damodaran ERP (India) | RBI 10-yr G-Sec | Motilal Oswal / ICICI Direct Research")
src.font = Font(name='Calibri', italic=True, color="595959", size=9)
src.alignment = al('left', indent=1)
cov.merge_cells('B54:I54')

# ─────────────────────────────────────────────────────────────────────────────
# TAB 2 — ASSUMPTIONS
# ─────────────────────────────────────────────────────────────────────────────
ass = wb.create_sheet("ASSUMP")
ass.sheet_view.showGridLines = False
set_widths(ass, {'A':38,'B':16,'C':55})

title_row(ass, 1, '  DIXON TECHNOLOGIES — MODEL ASSUMPTIONS  (Edit ONLY blue cells)', cols=3)

# ── SECTION 1: INCOME STATEMENT DRIVERS ──────────────────────────────────────
sub_hdr(ass, 3, '  INCOME STATEMENT DRIVERS', cols=3)
col_hdr(ass, 4, ['Assumption', 'Value / Rate', 'Notes / Source'], start_col=1, bg=MED_BLUE)

is_drivers = [
    # row, label,    value,   fmt,   note
    (5,  "Revenue Growth — FY26E",   0.30,  PCT,  "30% YoY; Dixon guided ₹50,000Cr+; PLI-driven mobile ramp-up"),
    (6,  "Revenue Growth — FY27E",   0.25,  PCT,  "25% YoY; continued EMS expansion + new category wins"),
    (7,  "Revenue Growth — FY28E",   0.18,  PCT,  "18% YoY; growth normalisation, base effect kicking in"),
    (8,  "Revenue Growth — FY29E",   0.15,  PCT,  "15% YoY; stable EMS market share gains"),
    (9,  "Revenue Growth — FY30E",   0.12,  PCT,  "12% YoY; mature growth phase, diversified revenue base"),
    (10, "EBITDA Margin % (all yrs)", 0.052, PCT,  "5.2%; slight margin expansion from operating leverage & PLI income"),
    (11, "D&A as % of Revenue",      0.008, PCT,  "0.8%; low asset-light model; capex mostly leased equipment"),
    (12, "Effective Tax Rate",        0.25,  PCT,  "25%; new tax regime post FY22; Source: Dixon Annual Reports"),
    (13, "Pre-tax Cost of Debt (Kd)", 0.085, PCT,  "8.5%; ~repo rate + 200bps; RBI repo at 6.5% as of June 2026"),
]
for row, label, val, fmt, note in is_drivers:
    lbl(ass, row, 1, label, bold=False, indent=1)
    inp(ass, row, 2, val, fmt=fmt, yellow=(row==10 or row==5))
    ass.cell(row=row, column=3).value = note
    ass.cell(row=row, column=3).font = Font(name='Calibri', color="595959", size=9, italic=True)
    ass.cell(row=row, column=3).alignment = al('left', indent=1)
    if row % 2 == 0:
        stripe(ass, row, cols=3, color=XLIGHT_BLUE)

# ── SECTION 2: BALANCE SHEET DRIVERS ─────────────────────────────────────────
sub_hdr(ass, 15, '  BALANCE SHEET & CASH FLOW DRIVERS', cols=3)
col_hdr(ass, 16, ['Assumption', 'Value / Rate', 'Notes / Source'], start_col=1, bg=MED_BLUE)

bs_drivers = [
    (17, "Capex as % of Revenue",       0.020, PCT, "2.0%; asset-light EMS; light capex relative to scale"),
    (18, "NWC as % of Revenue",         0.045, PCT, "4.5%; Dixon historically runs lean working capital cycles"),
    (19, "Shares Outstanding (lakhs)",  600,   INR, "6.0 crore shares = 600 lakh shares; Source: NSE/BSE filings"),
]
for row, label, val, fmt, note in bs_drivers:
    lbl(ass, row, 1, label, bold=False, indent=1)
    inp(ass, row, 2, val, fmt=fmt)
    ass.cell(row=row, column=3).value = note
    ass.cell(row=row, column=3).font = Font(name='Calibri', color="595959", size=9, italic=True)
    ass.cell(row=row, column=3).alignment = al('left', indent=1)
    if row % 2 == 1:
        stripe(ass, row, cols=3, color=XLIGHT_BLUE)

# ── SECTION 3: WACC INPUTS ────────────────────────────────────────────────────
sub_hdr(ass, 21, '  WACC / CAPM INPUTS  (Week 3 — WiDS)', cols=3)
col_hdr(ass, 22, ['Parameter', 'Value', 'Notes / Source'], start_col=1, bg=MED_BLUE)

wacc_inputs = [
    (24, "Risk-Free Rate (Rf)",         0.070, PCT, "7.0%; India 10-yr G-Sec yield (RBI, June 2026)"),
    (25, "Equity Risk Premium (ERP)",   0.060, PCT, "6.0%; Damodaran India ERP (Jan 2026 update)"),
    (26, "Beta (β) — Levered",          1.40,  PLAIN, "1.40; regressed vs. Nifty 50 (2-yr weekly); EMS sector cyclical"),
    (27, "Cost of Equity Ke = Rf+β×ERP","=B24+B26*B25", PCT, "CAPM: 7.0% + 1.40×6.0% = 15.4%"),
    (28, "Debt Weight D/(D+E)",         0.20,  PCT, "20%; conservative; Dixon is equity-heavy; Source: BS FY25"),
    (29, "Pre-tax Kd (gross)",          "=B13", PCT, "Linked from IS Drivers above"),
    (30, "After-tax Kd = Kd×(1-Tax)",  "=B29*(1-B12)", PCT, "8.5%×(1-25%) = 6.375%"),
    (31, "WACC = Ke×(1-D%) + Kd×D%",  "=B27*(1-B28)+B30*B28", PCT, "≈ 13.6%; weighted avg cost of capital"),
]
for row, label, val, fmt, note in wacc_inputs:
    lbl(ass, row, 1, label, bold=False, indent=1)
    if isinstance(val, str):
        c = ass.cell(row=row, column=2, value=val)
        c.font = Font(name='Calibri', color="000000", size=10)
        c.number_format = fmt
        c.alignment = al('right')
    else:
        inp(ass, row, 2, val, fmt=fmt, yellow=(row==31))
    ass.cell(row=row, column=3).value = note
    ass.cell(row=row, column=3).font = Font(name='Calibri', color="595959", size=9, italic=True)
    ass.cell(row=row, column=3).alignment = al('left', indent=1)
    if row % 2 == 0:
        stripe(ass, row, cols=3, color=XLIGHT_BLUE)

# ── SECTION 4: DCF TERMINAL VALUE ─────────────────────────────────────────────
sub_hdr(ass, 33, '  TERMINAL VALUE INPUT  (Week 3 — WiDS)', cols=3)
col_hdr(ass, 34, ['Parameter', 'Value', 'Notes / Source'], start_col=1, bg=MED_BLUE)
lbl(ass, 35, 1, "Terminal Growth Rate (g)", bold=False, indent=1)
inp(ass, 35, 2, 0.055, fmt=PCT, yellow=True)
ass.cell(row=35, column=3).value = "5.5%; long-run India nominal GDP growth proxy; conservative vs EMS sector"
ass.cell(row=35, column=3).font  = Font(name='Calibri', color="595959", size=9, italic=True)
ass.cell(row=35, column=3).alignment = al('left', indent=1)

# Source URLs block
sub_hdr(ass, 37, '  DATA SOURCES & LINKS', cols=3)
urls = [
    "Dixon AR FY25 (BSE):  https://www.bseindia.com/bseplus/AnnualReport/541988/10124541988.pdf",
    "Dixon AR FY24 (BSE):  https://www.bseindia.com/bseplus/AnnualReport/541988/10124541988.pdf",
    "Screener.in (Consolidated P&L/BS/CF):  https://www.screener.in/company/DIXON/consolidated/",
    "Trendlyne (financials + ratios):  https://trendlyne.com/equity/1624/DIXON/",
    "Damodaran ERP India:  https://pages.stern.nyu.edu/~adamodar/New_Home_Page/datafile/ctryprem.html",
    "RBI G-Sec yields:  https://www.rbi.org.in/Scripts/BS_NSDPDisplayGenericDetails.aspx?Id=219",
    "NSE India — DIXON:  https://www.nseindia.com/get-quotes/equity?symbol=DIXON",
]
for i, url in enumerate(urls, start=38):
    c = ass.cell(row=i, column=1, value=url)
    c.font = Font(name='Calibri', color="0070C0", size=9, italic=True, underline='single')
    c.alignment = al('left', indent=1)
    ass.merge_cells(start_row=i, start_column=1, end_row=i, end_column=3)

# ─────────────────────────────────────────────────────────────────────────────
# TAB 3 — INCOME STATEMENT  (+ Ratio Analysis)
# ─────────────────────────────────────────────────────────────────────────────
ws_is = wb.create_sheet("IS")
ws_is.sheet_view.showGridLines = False
set_widths(ws_is, {'A':32,'B':12,'C':12,'D':12,'E':12,'F':12,'G':12,'H':12,'I':12,'J':12})
ws_is.freeze_panes = 'B5'

title_row(ws_is, 1, '  DIXON TECHNOLOGIES — INCOME STATEMENT  (₹ Crore)', cols=10)
col_hdr(ws_is, 2, ['','FY22A','FY23A','FY24A','FY25A','FY26E','FY27E','FY28E','FY29E','FY30E'])
col_hdr(ws_is, 3, ['','Actual','Actual','Actual','Actual','Forecast','Forecast','Forecast','Forecast','Forecast'])

# ── REVENUE ───────────────────────────────────────────────────────────────────
sub_hdr(ws_is, 4, '  REVENUE', cols=10)

# Row 6: Revenue — historical = blue inputs, forecast = ASSUMP-driven formulas
lbl(ws_is, 6, 1, "Net Revenue / Net Sales", bold=True, indent=1)
rev_hist = [10697, 12192, 17691, 38860]
for ci, val in enumerate(rev_hist, start=2):
    inp(ws_is, 6, ci, val)
# Forecast revenue: prior_rev * (1 + growth_rate)
growth_rows = [5, 6, 7, 8, 9]   # ASSUMP rows for FY26-FY30 growth rates
for ci, grow_row in zip(range(6, 11), growth_rows):
    col = get_column_letter(ci)
    prev = get_column_letter(ci - 1)
    frm(ws_is, 6, ci, f'={prev}6*(1+ASSUMP!B{grow_row})')
stripe(ws_is, 6, cols=10, color=XLIGHT_BLUE)

# Row 7: YoY Growth %
lbl(ws_is, 7, 1, "  YoY Revenue Growth %", indent=2)
for ci in range(3, 11):
    col  = get_column_letter(ci)
    prev = get_column_letter(ci - 1)
    frm(ws_is, 7, ci, f'={col}6/{prev}6-1', fmt=PCT)
ws_is.cell(row=7, column=2).value = None

# ── COST STRUCTURE ────────────────────────────────────────────────────────────
sub_hdr(ws_is, 8, '  COST STRUCTURE', cols=10)

# Row 10: COGS/Raw Materials — hist=blue, fcst=derived (EBITDA% approach backsolve via EBITDA)
# We use a direct split: EBITDA = Revenue × EBITDA%; then EBITDA = Rev - COGS - Emp - Other
# For simplicity and WiDS level: COGS = Revenue × (1 - EBITDA% - Emp%_rev - Other%_rev)
# Hist ratios: Emp ~2.5%, Other ~5.5% → COGS is residual
lbl(ws_is, 10, 1, "Raw Materials / COGS", indent=2)
cogs_hist = [9330, 10630, 15440, 34000]
for ci, val in enumerate(cogs_hist, start=2):
    inp(ws_is, 10, ci, val)
for ci in range(6, 11):
    col = get_column_letter(ci)
    # COGS = Rev - EBITDA - Emp - Other; Emp ≈ 2.5% rev, Other ≈ 5.3% rev
    frm(ws_is, 10, ci, f'={col}6-(ASSUMP!B10*{col}6)-({col}6*0.025)-({col}6*0.053)')

# Row 11: Employee Costs
lbl(ws_is, 11, 1, "Employee Expenses", indent=2)
emp_hist = [307, 360, 490, 950]
for ci, val in enumerate(emp_hist, start=2):
    inp(ws_is, 11, ci, val)
for ci in range(6, 11):
    col = get_column_letter(ci)
    frm(ws_is, 11, ci, f'={col}6*0.025')

# Row 12: Other Opex
lbl(ws_is, 12, 1, "Other Operating Expenses", indent=2)
other_hist = [630, 692, 1053, 1967]
for ci, val in enumerate(other_hist, start=2):
    inp(ws_is, 12, ci, val)
for ci in range(6, 11):
    col = get_column_letter(ci)
    frm(ws_is, 12, ci, f'={col}6*0.053')

# ── EBITDA ────────────────────────────────────────────────────────────────────
# Note: row 13 is a DATA row — no merge here; styled via stripe + totals_row
lbl(ws_is, 13, 1, "EBITDA", bold=True, indent=1, bg=GREEN_FILL)
for ci in range(2, 11):
    col = get_column_letter(ci)
    frm(ws_is, 13, ci, f'={col}6-{col}10-{col}11-{col}12')
totals_row(ws_is, 13)

lbl(ws_is, 14, 1, "  EBITDA Margin %", indent=2)
for ci in range(2, 11):
    col = get_column_letter(ci)
    frm(ws_is, 14, ci, f'={col}13/{col}6', fmt=PCT)
stripe(ws_is, 14, color=XLIGHT_BLUE)

# ── DEPRECIATION & AMORTIZATION ───────────────────────────────────────────────
sub_hdr(ws_is, 15, '  BELOW EBITDA', cols=10)

lbl(ws_is, 17, 1, "Depreciation & Amortisation (D&A)", indent=2)
da_hist = [115, 140, 185, 280]
for ci, val in enumerate(da_hist, start=2):
    inp(ws_is, 17, ci, val)
for ci in range(6, 11):
    col = get_column_letter(ci)
    frm(ws_is, 17, ci, f'={col}6*ASSUMP!B11')

lbl(ws_is, 18, 1, "EBIT  (Operating Profit)", bold=True, indent=1)
for ci in range(2, 11):
    col = get_column_letter(ci)
    frm(ws_is, 18, ci, f'={col}13-{col}17')
totals_row(ws_is, 18)

lbl(ws_is, 19, 1, "  EBIT Margin %", indent=2)
for ci in range(2, 11):
    col = get_column_letter(ci)
    frm(ws_is, 19, ci, f'={col}18/{col}6', fmt=PCT)
stripe(ws_is, 19, color=XLIGHT_BLUE)

# ── BELOW EBIT ────────────────────────────────────────────────────────────────
lbl(ws_is, 22, 1, "Interest Expense", indent=2)
int_hist = [35, 45, 85, 120]
for ci, val in enumerate(int_hist, start=2):
    inp(ws_is, 22, ci, val)
for ci in range(6, 11):
    col = get_column_letter(ci)
    prev = get_column_letter(ci - 1)
    # Interest = prior year (LT Debt + ST Debt) × pre-tax Kd — consistent with WACC page (ASSUMP!B13)
    frm(ws_is, 22, ci, f'=ROUND((BS!{prev}21+BS!{prev}26)*ASSUMP!B13,0)', color="008000")

lbl(ws_is, 23, 1, "Profit Before Tax (PBT)", bold=True, indent=1)
for ci in range(2, 11):
    col = get_column_letter(ci)
    frm(ws_is, 23, ci, f'={col}18-{col}22')
totals_row(ws_is, 23)

lbl(ws_is, 24, 1, "Tax Expense", indent=2)
tax_hist = [68, 85, 108, 400]
for ci, val in enumerate(tax_hist, start=2):
    inp(ws_is, 24, ci, val)
for ci in range(6, 11):
    col = get_column_letter(ci)
    frm(ws_is, 24, ci, f'={col}23*ASSUMP!B12')

lbl(ws_is, 25, 1, "  Effective Tax Rate %", indent=2)
for ci in range(2, 11):
    col = get_column_letter(ci)
    frm(ws_is, 25, ci, f'={col}24/{col}23', fmt=PCT)
stripe(ws_is, 25, color=XLIGHT_BLUE)

lbl(ws_is, 26, 1, "Minority Interest (deduction)", indent=2)
mi_hist = [19, 0, -37, -61]
for ci, val in enumerate(mi_hist, start=2):
    inp(ws_is, 26, ci, val)
for ci in range(6, 11):
    col = get_column_letter(ci)
    frm(ws_is, 26, ci, f'=-ROUND({col}6*0.001,0)')

lbl(ws_is, 27, 1, "PAT  (Profit After Tax)", bold=True, indent=1)
for ci in range(2, 11):
    col = get_column_letter(ci)
    frm(ws_is, 27, ci, f'={col}23-{col}24-{col}26')
totals_row(ws_is, 27)

lbl(ws_is, 28, 1, "  PAT Margin %", indent=2)
for ci in range(2, 11):
    col = get_column_letter(ci)
    frm(ws_is, 28, ci, f'={col}27/{col}6', fmt=PCT)
stripe(ws_is, 28, color=XLIGHT_BLUE)

lbl(ws_is, 29, 1, "  EPS (₹ per share)", indent=2)
for ci in range(2, 11):
    col = get_column_letter(ci)
    frm(ws_is, 29, ci, f'={col}27*100/ASSUMP!B19', fmt=INR1)

# ════════════════════════════════════════════════════════════════════════════
# RATIO ANALYSIS SECTION  (WiDS Week 1)
# ════════════════════════════════════════════════════════════════════════════
ws_is.row_dimensions[31].height = 8

sub_hdr(ws_is, 32, '  RATIO ANALYSIS  ——  WiDS Week 1', bg=DARK_BLUE, cols=10)

# ── PROFITABILITY RATIOS ──────────────────────────────────────────────────────
sub_hdr(ws_is, 33, '  PROFITABILITY RATIOS', bg=MED_BLUE, cols=10)

lbl(ws_is, 34, 1, "Gross Margin  (Revenue − COGS) / Revenue", indent=2)
for ci in range(2, 11):
    col = get_column_letter(ci)
    frm(ws_is, 34, ci, f'=({col}6-{col}10)/{col}6', fmt=PCT)
stripe(ws_is, 34, color=XLIGHT_BLUE)

lbl(ws_is, 35, 1, "EBITDA Margin %", indent=2)
for ci in range(2, 11):
    col = get_column_letter(ci)
    frm(ws_is, 35, ci, f'={col}13/{col}6', fmt=PCT)

lbl(ws_is, 36, 1, "EBIT Margin %  (Operating Margin)", indent=2)
for ci in range(2, 11):
    col = get_column_letter(ci)
    frm(ws_is, 36, ci, f'={col}18/{col}6', fmt=PCT)
stripe(ws_is, 36, color=XLIGHT_BLUE)

lbl(ws_is, 37, 1, "PAT Margin %  (Net Profit Margin)", indent=2)
for ci in range(2, 11):
    col = get_column_letter(ci)
    frm(ws_is, 37, ci, f'={col}27/{col}6', fmt=PCT)

lbl(ws_is, 38, 1, "Return on Equity (ROE) = PAT / Equity", indent=2)
for ci in range(2, 11):
    col = get_column_letter(ci)
    frm(ws_is, 38, ci, f'={col}27/BS!{col}20', fmt=PCT, color="008000")
stripe(ws_is, 38, color=XLIGHT_BLUE)

lbl(ws_is, 39, 1, "Return on Assets (ROA) = PAT / Total Assets", indent=2)
for ci in range(2, 11):
    col = get_column_letter(ci)
    frm(ws_is, 39, ci, f'={col}27/BS!{col}17', fmt=PCT, color="008000")

lbl(ws_is, 40, 1, "ROCE = EBIT / (Total Assets − Current Liabilities)", indent=2)
for ci in range(2, 11):
    col = get_column_letter(ci)
    frm(ws_is, 40, ci, f'={col}18/(BS!{col}17-BS!{col}28)', fmt=PCT, color="008000")
stripe(ws_is, 40, color=XLIGHT_BLUE)

# ── EFFICIENCY RATIOS ─────────────────────────────────────────────────────────
sub_hdr(ws_is, 42, '  EFFICIENCY / ACTIVITY RATIOS  (days)', bg=MED_BLUE, cols=10)

lbl(ws_is, 43, 1, "Inventory Days  (Inventory / COGS × 365)", indent=2)
for ci in range(2, 11):
    col = get_column_letter(ci)
    frm(ws_is, 43, ci, f'=BS!{col}12/{col}10*365', fmt=PLAIN, color="008000")
stripe(ws_is, 43, color=XLIGHT_BLUE)

lbl(ws_is, 44, 1, "Receivable Days  (Trade Rec / Revenue × 365)", indent=2)
for ci in range(2, 11):
    col = get_column_letter(ci)
    frm(ws_is, 44, ci, f'=BS!{col}13/{col}6*365', fmt=PLAIN, color="008000")

lbl(ws_is, 45, 1, "Payable Days  (Trade Pay / COGS × 365)", indent=2)
for ci in range(2, 11):
    col = get_column_letter(ci)
    frm(ws_is, 45, ci, f'=BS!{col}25/{col}10*365', fmt=PLAIN, color="008000")
stripe(ws_is, 45, color=XLIGHT_BLUE)

lbl(ws_is, 46, 1, "Asset Turnover  (Revenue / Total Assets)", indent=2)
for ci in range(2, 11):
    col = get_column_letter(ci)
    frm(ws_is, 46, ci, f'={col}6/BS!{col}17', fmt=MULT, color="008000")

# ── LEVERAGE RATIOS ───────────────────────────────────────────────────────────
sub_hdr(ws_is, 48, '  LEVERAGE RATIOS', bg=MED_BLUE, cols=10)

lbl(ws_is, 49, 1, "Debt / Equity  ((LT Debt + ST Debt) / Equity)", indent=2)
for ci in range(2, 11):
    col = get_column_letter(ci)
    frm(ws_is, 49, ci, f'=(BS!{col}21+BS!{col}26)/BS!{col}20', fmt=MULT, color="008000")
stripe(ws_is, 49, color=XLIGHT_BLUE)

lbl(ws_is, 50, 1, "Interest Coverage  (EBIT / Interest)", indent=2)
for ci in range(2, 11):
    col = get_column_letter(ci)
    frm(ws_is, 50, ci, f'={col}18/{col}22', fmt=MULT)

lbl(ws_is, 51, 1, "Net Debt / EBITDA  ((Debt − Cash) / EBITDA)", indent=2)
for ci in range(2, 11):
    col = get_column_letter(ci)
    frm(ws_is, 51, ci, f'=(BS!{col}21+BS!{col}26-BS!{col}14)/{col}13', fmt=MULT, color="008000")
stripe(ws_is, 51, color=XLIGHT_BLUE)

# ─────────────────────────────────────────────────────────────────────────────
# TAB 4 — BALANCE SHEET
# ─────────────────────────────────────────────────────────────────────────────
ws_bs = wb.create_sheet("BS")
ws_bs.sheet_view.showGridLines = False
set_widths(ws_bs, {'A':32,'B':12,'C':12,'D':12,'E':12,'F':12,'G':12,'H':12,'I':12,'J':12})
ws_bs.freeze_panes = 'B5'

title_row(ws_bs, 1, '  DIXON TECHNOLOGIES — BALANCE SHEET  (₹ Crore)', cols=10)
col_hdr(ws_bs, 2, ['','FY22A','FY23A','FY24A','FY25A','FY26E','FY27E','FY28E','FY29E','FY30E'])
col_hdr(ws_bs, 3, ['','Actual','Actual','Actual','Actual','Forecast','Forecast','Forecast','Forecast','Forecast'])

# ── NON-CURRENT ASSETS ────────────────────────────────────────────────────────
sub_hdr(ws_bs, 4, '  NON-CURRENT ASSETS', cols=10)

lbl(ws_bs, 6, 1, "Net Fixed Assets (PP&E, net)", indent=2)
nfa_hist = [720, 850, 1150, 1800]
for ci, val in enumerate(nfa_hist, start=2):
    inp(ws_bs, 6, ci, val)
for ci in range(6, 11):
    col = get_column_letter(ci)
    prev = get_column_letter(ci - 1)
    # NFA = prev NFA + Capex - D&A
    frm(ws_bs, 6, ci, f'={prev}6+IS!{col}6*ASSUMP!B17-IS!{col}17', color="008000")

lbl(ws_bs, 7, 1, "Goodwill & Intangibles", indent=2)
gw_hist = [80, 75, 70, 65]
for ci, val in enumerate(gw_hist, start=2):
    inp(ws_bs, 7, ci, val)
for ci in range(6, 11):
    col = get_column_letter(ci)
    prev = get_column_letter(ci - 1)
    frm(ws_bs, 7, ci, f'={prev}7*0.95')   # slow amort

lbl(ws_bs, 8, 1, "Long-term Investments & Others", indent=2)
lt_hist = [120, 140, 200, 380]
for ci, val in enumerate(lt_hist, start=2):
    inp(ws_bs, 8, ci, val)
for ci in range(6, 11):
    col = get_column_letter(ci)
    prev = get_column_letter(ci - 1)
    frm(ws_bs, 8, ci, f'={prev}8*1.08')

lbl(ws_bs, 9, 1, "TOTAL NON-CURRENT ASSETS", bold=True, indent=1)
for ci in range(2, 11):
    col = get_column_letter(ci)
    frm(ws_bs, 9, ci, f'={col}6+{col}7+{col}8')
totals_row(ws_bs, 9)

# ── CURRENT ASSETS ────────────────────────────────────────────────────────────
sub_hdr(ws_bs, 10, '  CURRENT ASSETS', cols=10)

lbl(ws_bs, 12, 1, "Inventories", indent=2)
inv_hist = [750, 900, 1400, 2800]
for ci, val in enumerate(inv_hist, start=2):
    inp(ws_bs, 12, ci, val)
for ci in range(6, 11):
    col = get_column_letter(ci)
    frm(ws_bs, 12, ci, f'=IS!{col}6*0.072', color="008000")  # ~inventory days ~26

lbl(ws_bs, 13, 1, "Trade Receivables", indent=2)
rec_hist = [900, 1100, 1750, 4200]
for ci, val in enumerate(rec_hist, start=2):
    inp(ws_bs, 13, ci, val)
for ci in range(6, 11):
    col = get_column_letter(ci)
    frm(ws_bs, 13, ci, f'=IS!{col}6*0.108', color="008000")  # ~receivable days ~39

lbl(ws_bs, 14, 1, "Cash & Cash Equivalents", indent=2)
cash_hist = [320, 290, 450, 1200]
for ci, val in enumerate(cash_hist, start=2):
    inp(ws_bs, 14, ci, val)
for ci in range(6, 11):
    col = get_column_letter(ci)
    prev = get_column_letter(ci - 1)
    frm(ws_bs, 14, ci, f'={prev}14+CF!{col}22', color="008000")

lbl(ws_bs, 15, 1, "Other Current Assets", indent=2)
oca_hist = [450, 520, 750, 1500]
for ci, val in enumerate(oca_hist, start=2):
    inp(ws_bs, 15, ci, val)
for ci in range(6, 11):
    col = get_column_letter(ci)
    frm(ws_bs, 15, ci, f'=IS!{col}6*0.039', color="008000")

lbl(ws_bs, 16, 1, "TOTAL CURRENT ASSETS", bold=True, indent=1)
for ci in range(2, 11):
    col = get_column_letter(ci)
    frm(ws_bs, 16, ci, f'={col}12+{col}13+{col}14+{col}15')
totals_row(ws_bs, 16)

lbl(ws_bs, 17, 1, "TOTAL ASSETS", bold=True, indent=1)
for ci in range(2, 11):
    col = get_column_letter(ci)
    frm(ws_bs, 17, ci, f'={col}9+{col}16')
for ci in range(2, 11):
    ws_bs.cell(row=17, column=ci).font = Font(name='Calibri', bold=True, color="000000", size=10)
stripe(ws_bs, 17, color=LIGHT_BLUE)

# ── EQUITY ────────────────────────────────────────────────────────────────────
sub_hdr(ws_bs, 18, '  EQUITY & LIABILITIES', cols=10)

lbl(ws_bs, 20, 1, "Shareholders' Equity", bold=True, indent=1)
eq_hist = [1100, 1350, 1720, 2940]
for ci, val in enumerate(eq_hist, start=2):
    inp(ws_bs, 20, ci, val)
for ci in range(6, 11):
    col = get_column_letter(ci)
    prev = get_column_letter(ci - 1)
    frm(ws_bs, 20, ci, f'={prev}20+IS!{col}27+CF!{col}19', color="008000")  # +CF div (negative) deducts dividends

lbl(ws_bs, 21, 1, "Long-term Borrowings (LT Debt)", indent=2)
ltd_hist = [180, 220, 350, 600]
for ci, val in enumerate(ltd_hist, start=2):
    inp(ws_bs, 21, ci, val)
for ci in range(6, 11):
    col = get_column_letter(ci)
    prev = get_column_letter(ci - 1)
    frm(ws_bs, 21, ci, f'={prev}21*1.05')

lbl(ws_bs, 22, 1, "Deferred Tax & Other NCL", indent=2)
dtl_hist = [80, 90, 120, 180]
for ci, val in enumerate(dtl_hist, start=2):
    inp(ws_bs, 22, ci, val)
for ci in range(6, 11):
    col = get_column_letter(ci)
    prev = get_column_letter(ci - 1)
    frm(ws_bs, 22, ci, f'={prev}22*1.04')

# ── CURRENT LIABILITIES ───────────────────────────────────────────────────────
sub_hdr(ws_bs, 23, '  CURRENT LIABILITIES', cols=10)

lbl(ws_bs, 25, 1, "Trade Payables", indent=2)
tp_hist = [900, 1050, 1600, 3800]
for ci, val in enumerate(tp_hist, start=2):
    inp(ws_bs, 25, ci, val)
for ci in range(6, 11):
    col = get_column_letter(ci)
    frm(ws_bs, 25, ci, f'=IS!{col}10*0.112', color="008000")  # ~payable days ~41

lbl(ws_bs, 26, 1, "Short-term Borrowings (ST Debt)", indent=2)
std_hist = [250, 300, 420, 500]
for ci, val in enumerate(std_hist, start=2):
    inp(ws_bs, 26, ci, val)
for ci in range(6, 11):
    col = get_column_letter(ci)
    prev = get_column_letter(ci - 1)
    frm(ws_bs, 26, ci, f'={prev}26*0.95')

lbl(ws_bs, 27, 1, "Other Current Liabilities", indent=2)
ocl_hist = [600, 720, 1100, 2600]
for ci, val in enumerate(ocl_hist, start=2):
    inp(ws_bs, 27, ci, val)
for ci in range(6, 11):
    col = get_column_letter(ci)
    frm(ws_bs, 27, ci, f'=IS!{col}6*0.067', color="008000")

lbl(ws_bs, 28, 1, "TOTAL CURRENT LIABILITIES", bold=True, indent=1)
for ci in range(2, 11):
    col = get_column_letter(ci)
    frm(ws_bs, 28, ci, f'={col}25+{col}26+{col}27')
totals_row(ws_bs, 28)

lbl(ws_bs, 29, 1, "TOTAL EQUITY & LIABILITIES", bold=True, indent=1)
for ci in range(2, 11):
    col = get_column_letter(ci)
    frm(ws_bs, 29, ci, f'={col}20+{col}21+{col}22+{col}28')
stripe(ws_bs, 29, color=LIGHT_BLUE)
for ci in range(2, 11):
    ws_bs.cell(row=29, column=ci).font = Font(name='Calibri', bold=True, size=10)

# Row 30: Balance Check
lbl(ws_bs, 30, 1, "  ✔ Balance Check  (Assets − L&E = 0?)", indent=2)
for ci in range(2, 11):
    col = get_column_letter(ci)
    c = frm(ws_bs, 30, ci, f'={col}17-{col}29')
    c.font = Font(name='Calibri', bold=True, color="FF0000", size=10)

# ─────────────────────────────────────────────────────────────────────────────
# TAB 5 — CASH FLOW STATEMENT
# ─────────────────────────────────────────────────────────────────────────────
ws_cf = wb.create_sheet("CF")
ws_cf.sheet_view.showGridLines = False
set_widths(ws_cf, {'A':35,'B':12,'C':12,'D':12,'E':12,'F':12,'G':12,'H':12,'I':12,'J':12})
ws_cf.freeze_panes = 'B5'

title_row(ws_cf, 1, '  DIXON TECHNOLOGIES — CASH FLOW STATEMENT  (₹ Crore, Indirect Method)', cols=10)
col_hdr(ws_cf, 2, ['','FY22A','FY23A','FY24A','FY25A','FY26E','FY27E','FY28E','FY29E','FY30E'])
col_hdr(ws_cf, 3, ['','Actual','Actual','Actual','Actual','Forecast','Forecast','Forecast','Forecast','Forecast'])

# ── OPERATING ACTIVITIES ──────────────────────────────────────────────────────
sub_hdr(ws_cf, 4, '  OPERATING ACTIVITIES', cols=10)

lbl(ws_cf, 5, 1, "Net Profit (PAT)", bold=True, indent=1)
for ci in range(2, 11):
    col = get_column_letter(ci)
    frm(ws_cf, 5, ci, f'=IS!{col}27', color="008000")
stripe(ws_cf, 5, color=XLIGHT_BLUE)

lbl(ws_cf, 6, 1, "  Add: D&A (Non-cash)", indent=2)
for ci in range(2, 11):
    col = get_column_letter(ci)
    frm(ws_cf, 6, ci, f'=IS!{col}17', color="008000")

lbl(ws_cf, 7, 1, "  Changes in Working Capital (ΔNWC)", indent=2)
# ΔNWC = −(NWC_curr - NWC_prev); NWC = (Inv + Rec + OCA) - (TP + OCL)
# For historical: actual change, for forecast: ASSUMP!B18 × Revenue change
nwc_hist = [-120, -180, -320, -950]  # approximate hist WC changes
for ci, val in enumerate(nwc_hist, start=2):
    inp(ws_cf, 7, ci, val)
for ci in range(6, 11):
    col = get_column_letter(ci)
    prev = get_column_letter(ci - 1)
    frm(ws_cf, 7, ci, f'=-(IS!{col}6-IS!{prev}6)*ASSUMP!B18', color="008000")
stripe(ws_cf, 7, color=XLIGHT_BLUE)

lbl(ws_cf, 8, 1, "  Other Adjustments (Interest & Tax paid)", indent=2)
adj_hist = [-30, -40, -75, -110]
for ci, val in enumerate(adj_hist, start=2):
    inp(ws_cf, 8, ci, val)
for ci in range(6, 11):
    col = get_column_letter(ci)
    frm(ws_cf, 8, ci, f'=-(IS!{col}22+IS!{col}24)*0.05')

lbl(ws_cf, 9, 1, "CASH FROM OPERATIONS (CFO)", bold=True, indent=1)
for ci in range(2, 11):
    col = get_column_letter(ci)
    frm(ws_cf, 9, ci, f'={col}5+{col}6+{col}7+{col}8')
totals_row(ws_cf, 9)
stripe(ws_cf, 9, color=LIGHT_BLUE)

# ── INVESTING ACTIVITIES ──────────────────────────────────────────────────────
sub_hdr(ws_cf, 11, '  INVESTING ACTIVITIES', cols=10)

lbl(ws_cf, 12, 1, "  Capital Expenditure (Capex)", indent=2)
capex_hist = [-155, -195, -340, -600]
for ci, val in enumerate(capex_hist, start=2):
    inp(ws_cf, 12, ci, val)
for ci in range(6, 11):
    col = get_column_letter(ci)
    frm(ws_cf, 12, ci, f'=-(IS!{col}6*ASSUMP!B17)', color="008000")

lbl(ws_cf, 13, 1, "  Investments & Others", indent=2)
inv_cf_hist = [-40, -25, -65, -180]
for ci, val in enumerate(inv_cf_hist, start=2):
    inp(ws_cf, 13, ci, val)
for ci in range(6, 11):
    col = get_column_letter(ci)
    frm(ws_cf, 13, ci, f'=-IS!{col}6*0.005')

lbl(ws_cf, 14, 1, "CASH FROM INVESTING (CFI)", bold=True, indent=1)
for ci in range(2, 11):
    col = get_column_letter(ci)
    frm(ws_cf, 14, ci, f'={col}12+{col}13')
totals_row(ws_cf, 14)

# ── FINANCING ACTIVITIES ──────────────────────────────────────────────────────
sub_hdr(ws_cf, 16, '  FINANCING ACTIVITIES', cols=10)

lbl(ws_cf, 17, 1, "  Proceeds from / (Repayment of) Debt", indent=2)
debt_cf_hist = [80, 60, 120, 80]
for ci, val in enumerate(debt_cf_hist, start=2):
    inp(ws_cf, 17, ci, val)
for ci in range(6, 11):
    col = get_column_letter(ci)
    frm(ws_cf, 17, ci, f'=IS!{col}6*0.002')

lbl(ws_cf, 18, 1, "  Interest Paid", indent=2)
for ci in range(2, 11):
    col = get_column_letter(ci)
    frm(ws_cf, 18, ci, f'=-IS!{col}22', color="008000")
stripe(ws_cf, 18, color=XLIGHT_BLUE)

lbl(ws_cf, 19, 1, "  Dividends Paid", indent=2)
div_hist = [-18, -22, -28, -60]
for ci, val in enumerate(div_hist, start=2):
    inp(ws_cf, 19, ci, val)
for ci in range(6, 11):
    col = get_column_letter(ci)
    frm(ws_cf, 19, ci, f'=-IS!{col}27*0.08')  # ~8% dividend payout

lbl(ws_cf, 20, 1, "CASH FROM FINANCING (CFF)", bold=True, indent=1)
for ci in range(2, 11):
    col = get_column_letter(ci)
    frm(ws_cf, 20, ci, f'={col}17+{col}18+{col}19')
totals_row(ws_cf, 20)

# ── NET CHANGE IN CASH ────────────────────────────────────────────────────────
ws_cf.row_dimensions[21].height = 4
lbl(ws_cf, 22, 1, "NET CHANGE IN CASH  (CFO + CFI + CFF)", bold=True, indent=1)
for ci in range(2, 11):
    col = get_column_letter(ci)
    frm(ws_cf, 22, ci, f'={col}9+{col}14+{col}20')
stripe(ws_cf, 22, color=LIGHT_BLUE)
for ci in range(2, 11):
    ws_cf.cell(row=22, column=ci).font = Font(name='Calibri', bold=True, size=10)

lbl(ws_cf, 23, 1, "  Opening Cash Balance", indent=2)
open_cash = [None, 320, 290, 450]
for ci, val in enumerate(open_cash, start=2):
    if val is not None:
        inp(ws_cf, 23, ci, val)
for ci in range(6, 11):
    col = get_column_letter(ci)
    prev = get_column_letter(ci - 1)
    frm(ws_cf, 23, ci, f'=BS!{prev}14', color="008000")
# FY26 opening = FY25 closing
frm(ws_cf, 23, 6, f'=BS!E14', color="008000")

lbl(ws_cf, 24, 1, "  Closing Cash Balance", indent=2)
for ci in range(2, 11):
    col = get_column_letter(ci)
    frm(ws_cf, 24, ci, f'={col}23+{col}22')
stripe(ws_cf, 24, color=XLIGHT_BLUE)

# ─────────────────────────────────────────────────────────────────────────────
# TAB 6 — DCF  (WACC merged at top)
# ─────────────────────────────────────────────────────────────────────────────
ws_dcf = wb.create_sheet("DCF")
ws_dcf.sheet_view.showGridLines = False
set_widths(ws_dcf, {'A':35,'B':16,'C':14,'D':14,'E':14,'F':14,'G':14,'H':14})
ws_dcf.freeze_panes = 'B4'

title_row(ws_dcf, 1, '  DIXON TECHNOLOGIES — DCF VALUATION  +  WACC  (₹ Crore)', cols=8)

# ════════════════════════════════════════════════════════════════════════════
# SECTION A — WACC DERIVATION
# ════════════════════════════════════════════════════════════════════════════
sub_hdr(ws_dcf, 3, '  SECTION A — WACC DERIVATION  (WiDS Week 3 — CAPM)', cols=8)
col_hdr(ws_dcf, 4, ['Parameter', 'Value', 'Notes'], start_col=1, bg=MED_BLUE)
ws_dcf.merge_cells(start_row=4, start_column=3, end_row=4, end_column=8)

wacc_display = [
    (5,  "Risk-Free Rate (Rf)",            "=ASSUMP!B24", PCT,  "India 10-yr G-Sec yield — RBI (June 2026)"),
    (6,  "Equity Risk Premium (ERP)",       "=ASSUMP!B25", PCT,  "Damodaran India ERP (Jan 2026)"),
    (7,  "Beta (β)  — Levered vs Nifty 50", "=ASSUMP!B26", PLAIN,"Regressed 2-yr weekly; EMS is cyclical"),
    (8,  "Cost of Equity  Ke = Rf + β×ERP", "=ASSUMP!B27", PCT,  "CAPM: 7.0% + 1.40 × 6.0% = 15.4%"),
    (9,  "Weight of Debt  D/(D+E)",         "=ASSUMP!B28", PCT,  "Conservative; Dixon equity-heavy per FY25 BS"),
    (10, "Pre-tax Cost of Debt  Kd",        "=ASSUMP!B29", PCT,  "Repo + ~200bps spread"),
    (11, "After-tax Cost of Debt  Kd×(1-t)","=ASSUMP!B30", PCT,  "Tax shield on interest"),
    (12, "Weight of Equity  E/(D+E)",       "=1-ASSUMP!B28",PCT, "Complement of debt weight"),
    (13, "Terminal Growth Rate (g)",        "=ASSUMP!B35", PCT,  "Long-run India nominal GDP growth proxy"),
    (15, "► WACC  =  Ke×(E%) + Kd×(D%)",  "=ASSUMP!B31", PCT,  "≈ 13.6%  |  Ke×80% + Kd_at×20%"),
]
for row, param, val, fmt, note in wacc_display:
    lbl(ws_dcf, row, 1, param, indent=1)
    c = ws_dcf.cell(row=row, column=2, value=val)
    c.font = Font(name='Calibri', color="008000", size=10)
    c.number_format = fmt
    c.alignment = al('right')
    ws_dcf.merge_cells(start_row=row, start_column=3, end_row=row, end_column=8)
    nc = ws_dcf.cell(row=row, column=3, value=note)
    nc.font = Font(name='Calibri', color="595959", size=9, italic=True)
    nc.alignment = al('left', indent=1)
    if row % 2 == 0:
        stripe(ws_dcf, row, cols=8, color=XLIGHT_BLUE)

# Highlight WACC row
stripe(ws_dcf, 15, cols=8, color=LIGHT_YELLOW)
ws_dcf.cell(row=15, column=1).font = Font(name='Calibri', bold=True, size=10)
ws_dcf.cell(row=15, column=2).font = Font(name='Calibri', bold=True, color="008000", size=11)
ws_dcf.row_dimensions[15].height = 18

# ════════════════════════════════════════════════════════════════════════════
# SECTION B — FREE CASH FLOW TO FIRM (FCFF)
# ════════════════════════════════════════════════════════════════════════════
ws_dcf.row_dimensions[17].height = 6
sub_hdr(ws_dcf, 18, '  SECTION B — FREE CASH FLOW TO FIRM (FCFF)  FY26E–FY30E', cols=8)
col_hdr(ws_dcf, 19, ['', '', 'FY26E', 'FY27E', 'FY28E', 'FY29E', 'FY30E', ''], start_col=1, bg=MED_BLUE)

# fcff_cols: (IS_col_letter, DCF_col_index)
fcff_cols = [('F', 3), ('G', 4), ('H', 5), ('I', 6), ('J', 7)]

fcff_rows = [
    (20, "Revenue",               lambda ltr: f'=IS!{ltr}6'),
    (21, "EBIT",                  lambda ltr: f'=IS!{ltr}18'),
    (22, "EBIT Margin %",         lambda ltr: f'=IS!{ltr}19'),
    (23, "Tax Rate",              lambda ltr: f'=ASSUMP!B12'),
    (24, "NOPAT  = EBIT × (1-t)", lambda ltr: f'=DCF!{get_column_letter(fcff_cols[[x[0] for x in fcff_cols].index(ltr)][1])}21*(1-ASSUMP!B12)'),
    (25, "Add: D&A",              lambda ltr: f'=IS!{ltr}17'),
    (26, "Less: Capex",           lambda ltr: f'=-(IS!{ltr}6*ASSUMP!B17)'),
    (27, "Less: ΔNWC",            lambda ltr: f'=-(IS!{ltr}6*ASSUMP!B18-IS!{chr(ord(ltr)-1)}6*ASSUMP!B18)'),
]

# Build NOPAT mapping properly
lbl(ws_dcf, 20, 1, "Revenue (₹ Cr)", indent=1)
lbl(ws_dcf, 21, 1, "EBIT (₹ Cr)", indent=1)
lbl(ws_dcf, 22, 1, "  EBIT Margin %", indent=2)
lbl(ws_dcf, 23, 1, "  Tax Rate", indent=2)
lbl(ws_dcf, 24, 1, "NOPAT  =  EBIT × (1 − Tax Rate)", bold=True, indent=1)
lbl(ws_dcf, 25, 1, "  Add: D&A", indent=2)
lbl(ws_dcf, 26, 1, "  Less: Capex", indent=2)
lbl(ws_dcf, 27, 1, "FCFF  =  NOPAT + D&A − Capex − ΔNWC", bold=True, indent=1)
lbl(ws_dcf, 28, 1, "  Less: ΔNWC  (Change in Net Working Capital)", indent=2)

for ltr, dc in fcff_cols:
    dc_col = get_column_letter(dc)
    prev_ltr = chr(ord(ltr) - 1)
    frm(ws_dcf, 20, dc, f'=IS!{ltr}6', color="008000")
    frm(ws_dcf, 21, dc, f'=IS!{ltr}18', color="008000")
    frm(ws_dcf, 22, dc, f'=IS!{ltr}19', fmt=PCT, color="008000")
    inp(ws_dcf, 23, dc, None, fmt=PCT)
    ws_dcf.cell(row=23, column=dc).value = '=ASSUMP!B12'
    ws_dcf.cell(row=23, column=dc).font = Font(name='Calibri', color="008000", size=10)
    ws_dcf.cell(row=23, column=dc).number_format = PCT
    frm(ws_dcf, 24, dc, f'={dc_col}21*(1-ASSUMP!B12)')
    frm(ws_dcf, 25, dc, f'=IS!{ltr}17', color="008000")
    frm(ws_dcf, 26, dc, f'=-(IS!{ltr}6*ASSUMP!B17)')
    # ΔNWC = NWC_curr - NWC_prev
    frm(ws_dcf, 28, dc, f'=-(IS!{ltr}6*ASSUMP!B18-IS!{prev_ltr}6*ASSUMP!B18)')
    # FCFF = NOPAT + DA - Capex + ΔNWC (ΔNWC already negative if NWC rises)
    frm(ws_dcf, 27, dc, f'={dc_col}24+{dc_col}25+{dc_col}26+{dc_col}28')

stripe(ws_dcf, 27, cols=8, color=LIGHT_BLUE)
totals_row(ws_dcf, 27)
for dc in range(3, 8):
    ws_dcf.cell(row=27, column=dc).font = Font(name='Calibri', bold=True, size=10)
stripe(ws_dcf, 22, cols=8, color=XLIGHT_BLUE)
stripe(ws_dcf, 23, cols=8, color=XLIGHT_BLUE)

# ════════════════════════════════════════════════════════════════════════════
# SECTION C — PRESENT VALUE OF FCFF
# ════════════════════════════════════════════════════════════════════════════
ws_dcf.row_dimensions[30].height = 6
sub_hdr(ws_dcf, 31, '  SECTION C — PRESENT VALUE OF FCFF', cols=8)
col_hdr(ws_dcf, 32, ['', '', 'FY26E', 'FY27E', 'FY28E', 'FY29E', 'FY30E', ''], start_col=1, bg=MED_BLUE)

lbl(ws_dcf, 33, 1, "Discount Period (t)", indent=1)
lbl(ws_dcf, 34, 1, "Discount Factor  =  1 / (1+WACC)^t", indent=1)
lbl(ws_dcf, 35, 1, "PV of FCFF", bold=True, indent=1)

for i, (ltr, dc) in enumerate(fcff_cols, start=1):
    dc_col = get_column_letter(dc)
    inp(ws_dcf, 33, dc, i, fmt=PLAIN)
    frm(ws_dcf, 34, dc, f'=1/(1+ASSUMP!B31)^{dc_col}33', fmt='0.0000')
    frm(ws_dcf, 35, dc, f'={dc_col}27*{dc_col}34')

stripe(ws_dcf, 34, cols=8, color=XLIGHT_BLUE)
stripe(ws_dcf, 35, cols=8, color=LIGHT_BLUE)
totals_row(ws_dcf, 35)

# ════════════════════════════════════════════════════════════════════════════
# SECTION D — TERMINAL VALUE
# ════════════════════════════════════════════════════════════════════════════
ws_dcf.row_dimensions[37].height = 6
sub_hdr(ws_dcf, 38, '  SECTION D — TERMINAL VALUE  (Gordon Growth Model)', cols=8)

lbl(ws_dcf, 39, 1, "Terminal Year FCFF  (FY30E)", indent=1)
frm(ws_dcf, 39, 3, '=G27', color="008000")
ws_dcf.merge_cells(start_row=39, start_column=3, end_row=39, end_column=7)

lbl(ws_dcf, 40, 1, "Terminal Growth Rate  (g)", indent=1)
frm(ws_dcf, 40, 3, '=ASSUMP!B35', fmt=PCT, color="008000")

lbl(ws_dcf, 41, 1, "WACC", indent=1)
frm(ws_dcf, 41, 3, '=ASSUMP!B31', fmt=PCT, color="008000")

lbl(ws_dcf, 42, 1, "Terminal Value  =  FCFF×(1+g) / (WACC − g)", bold=True, indent=1)
frm(ws_dcf, 42, 3, '=C39*(1+ASSUMP!B35)/(ASSUMP!B31-ASSUMP!B35)')
stripe(ws_dcf, 42, cols=8, color=LIGHT_YELLOW)

lbl(ws_dcf, 43, 1, "PV of Terminal Value  =  TV / (1+WACC)^5", bold=True, indent=1)
frm(ws_dcf, 43, 3, '=C42/(1+ASSUMP!B31)^5')
stripe(ws_dcf, 43, cols=8, color=LIGHT_BLUE)
totals_row(ws_dcf, 43)

# ════════════════════════════════════════════════════════════════════════════
# SECTION E — ENTERPRISE VALUE BRIDGE
# ════════════════════════════════════════════════════════════════════════════
ws_dcf.row_dimensions[45].height = 6
sub_hdr(ws_dcf, 46, '  SECTION E — ENTERPRISE VALUE → EQUITY VALUE BRIDGE', cols=8)

lbl(ws_dcf, 47, 1, "Sum of PV (FCFF)  — FY26–FY30", indent=1)
frm(ws_dcf, 47, 3, '=SUM(C35:G35)')

lbl(ws_dcf, 48, 1, "PV of Terminal Value", indent=1)
frm(ws_dcf, 48, 3, '=C43', color="008000")
stripe(ws_dcf, 48, color=XLIGHT_BLUE)

lbl(ws_dcf, 49, 1, "Enterprise Value (EV)", bold=True, indent=1)
frm(ws_dcf, 49, 3, '=C47+C48')
stripe(ws_dcf, 49, cols=8, color=LIGHT_BLUE)
ws_dcf.cell(row=49, column=3).font = Font(name='Calibri', bold=True, size=11)
totals_row(ws_dcf, 49)

lbl(ws_dcf, 51, 1, "  Less: Total Debt  (LT + ST)", indent=2)
# Debt from BS FY25 (last actual) — blue input; update when forecast debt is modelled
inp(ws_dcf, 51, 3, None)
ws_dcf.cell(row=51, column=3).value = '=-(BS!E21+BS!E26)'
ws_dcf.cell(row=51, column=3).font = Font(name='Calibri', color="008000", size=10)
ws_dcf.cell(row=51, column=3).number_format = INR

lbl(ws_dcf, 52, 1, "  Add: Cash & Equivalents  (FY25A)", indent=2)
ws_dcf.cell(row=52, column=3).value = '=BS!E14'
ws_dcf.cell(row=52, column=3).font = Font(name='Calibri', color="008000", size=10)
ws_dcf.cell(row=52, column=3).number_format = INR
stripe(ws_dcf, 52, color=XLIGHT_BLUE)

lbl(ws_dcf, 53, 1, "  Less: Minority Interest  (FY25A)", indent=2)
ws_dcf.cell(row=53, column=3).value = '=-100'
ws_dcf.cell(row=53, column=3).font = Font(name='Calibri', color="0070C0", size=10)
ws_dcf.cell(row=53, column=3).number_format = INR

lbl(ws_dcf, 54, 1, "EQUITY VALUE  =  EV − Debt + Cash − MI", bold=True, indent=1)
frm(ws_dcf, 54, 3, '=C49+C51+C52+C53')
stripe(ws_dcf, 54, cols=8, color=LIGHT_BLUE)
totals_row(ws_dcf, 54)

lbl(ws_dcf, 55, 1, "  Shares Outstanding  (lakhs)", indent=2)
frm(ws_dcf, 55, 3, '=ASSUMP!B19', color="008000")

lbl(ws_dcf, 56, 1, "IMPLIED SHARE PRICE  (₹ per share)", bold=True, indent=1)
frm(ws_dcf, 56, 3, '=C54*100/ASSUMP!B19', fmt=INR)
stripe(ws_dcf, 56, cols=8, color=LIGHT_YELLOW)
ws_dcf.cell(row=56, column=1).font = Font(name='Calibri', bold=True, color=DARK_BLUE, size=11)
ws_dcf.cell(row=56, column=3).font = Font(name='Calibri', bold=True, color=DARK_BLUE, size=12)
ws_dcf.row_dimensions[56].height = 20

# ════════════════════════════════════════════════════════════════════════════
# SECTION F — EV/EBITDA CROSS-CHECK
# ════════════════════════════════════════════════════════════════════════════
ws_dcf.row_dimensions[58].height = 6
sub_hdr(ws_dcf, 59, '  SECTION F — EV/EBITDA CROSS-CHECK  (Sanity Test)', cols=8)

lbl(ws_dcf, 60, 1, "Enterprise Value (₹ Cr)", indent=1)
frm(ws_dcf, 60, 3, '=C49', color="008000")

lbl(ws_dcf, 61, 1, "LTM EBITDA  (FY25A, ₹ Cr)", indent=1)
frm(ws_dcf, 61, 3, '=IS!E13', color="008000")
stripe(ws_dcf, 61, color=XLIGHT_BLUE)

lbl(ws_dcf, 62, 1, "Implied EV/EBITDA Multiple", bold=True, indent=1)
frm(ws_dcf, 62, 3, '=C60/C61', fmt=MULT)
stripe(ws_dcf, 62, cols=8, color=LIGHT_YELLOW)

lbl(ws_dcf, 63, 1, "Peer Median EV/EBITDA  (see COMPS tab)", indent=1)
inp(ws_dcf, 63, 3, 42.6, fmt=MULT)
ws_dcf.cell(row=63, column=4, value="→ Check COMPS tab for peer context")
ws_dcf.cell(row=63, column=4).font = Font(name='Calibri', color="595959", size=9, italic=True)

# ─────────────────────────────────────────────────────────────────────────────
# TAB 7 — SENSITIVITY ANALYSIS
# ─────────────────────────────────────────────────────────────────────────────
ws_sens = wb.create_sheet("SENSITIVITY")
ws_sens.sheet_view.showGridLines = False
set_widths(ws_sens, {'A':5,'B':18,'C':14,'D':14,'E':14,'F':14,'G':14,'H':14,'I':14,'J':5})

title_row(ws_sens, 1, '  DIXON TECHNOLOGIES — 2-WAY SENSITIVITY  |  Implied Share Price (₹)', cols=10)

ws_sens.merge_cells('B3:J3')
info = ws_sens.cell(row=3, column=2,
    value="Implied Share Price (₹) sensitised across WACC (rows) and Terminal Growth Rate (columns). "
          "Formula replicates full DCF inline — no circular references.")
info.font = Font(name='Calibri', italic=True, color="595959", size=9)
info.alignment = al('left', indent=1)

# WACC axis: 11% to 16% in 1% steps
wacc_vals = [0.11, 0.12, 0.13, 0.136, 0.14, 0.15, 0.16]
# TGR axis: 4% to 7% in 0.5% steps
tgr_vals  = [0.040, 0.045, 0.050, 0.055, 0.060, 0.065, 0.070]

# Column headers (TGR)
ws_sens.cell(row=5, column=2, value="WACC  ↓  /  TGR  →").font = Font(name='Calibri', bold=True, color="FFFFFF", size=10)
ws_sens.cell(row=5, column=2).fill = fill(DARK_BLUE)
ws_sens.cell(row=5, column=2).alignment = al('center')
for j, tgr in enumerate(tgr_vals, start=3):
    c = ws_sens.cell(row=5, column=j, value=tgr)
    c.font = Font(name='Calibri', bold=True, color="FFFFFF", size=10)
    c.fill = fill(MED_BLUE)
    c.number_format = PCT
    c.alignment = al('center')

# Row headers (WACC) + body formulas
for i, wacc in enumerate(wacc_vals, start=6):
    # Row header
    c = ws_sens.cell(row=i, column=2, value=wacc)
    c.font = Font(name='Calibri', bold=True, color="FFFFFF", size=10)
    c.fill = fill(MED_BLUE)
    c.number_format = PCT
    c.alignment = al('center')

    for j, tgr in enumerate(tgr_vals, start=3):
        wacc_ref = get_column_letter(2) + str(i)   # B6..B12
        tgr_ref  = get_column_letter(j) + "5"       # C5..I5

        # Full inline DCF formula:
        # SUM of PV(FCFF yr1..5) + PV(TV) - Debt + Cash - MI, all ÷ shares × 100
        # FCFF at DCF!C27..G27; Debt @ DCF!C51; Cash @ DCF!C52; MI @ DCF!C53; Shares ASSUMP!B19
        formula = (
            f'=IF({tgr_ref}>={wacc_ref},"N/A",'
            f'(DCF!C27/(1+{wacc_ref})^1'
            f'+DCF!D27/(1+{wacc_ref})^2'
            f'+DCF!E27/(1+{wacc_ref})^3'
            f'+DCF!F27/(1+{wacc_ref})^4'
            f'+DCF!G27/(1+{wacc_ref})^5'
            f'+DCF!G27*(1+{tgr_ref})/({wacc_ref}-{tgr_ref})/(1+{wacc_ref})^5'
            f'+DCF!C51+DCF!C52+DCF!C53)'
            f'*100/ASSUMP!B19)'
        )
        cell = ws_sens.cell(row=i, column=j, value=formula)
        cell.number_format = INR
        cell.alignment = al('center')
        cell.font = Font(name='Calibri', size=10)

        # Highlight base case (≈WACC 13.6%, TGR 5.5%)
        if abs(wacc - 0.136) < 0.001 and abs(tgr - 0.055) < 0.001:
            cell.fill = fill(LIGHT_YELLOW)
            cell.font = Font(name='Calibri', bold=True, size=10)
        elif i % 2 == 0:
            cell.fill = fill(XLIGHT_BLUE)

ws_sens.row_dimensions[14].height = 8
note = ws_sens.cell(row=15, column=2,
    value="★  Yellow cell = Base Case (WACC=13.6%, TGR=5.5%).  "
          "N/A = mathematically undefined when TGR ≥ WACC (Gordon Growth Model constraint).")
note.font = Font(name='Calibri', italic=True, color="595959", size=9)
note.alignment = al('left', indent=1)
ws_sens.merge_cells('B15:J15')

# ─────────────────────────────────────────────────────────────────────────────
# TAB 8 — PEER COMPARABLES  (WiDS Week 4)
# ─────────────────────────────────────────────────────────────────────────────
ws_comp = wb.create_sheet("COMPS")
ws_comp.sheet_view.showGridLines = False
set_widths(ws_comp, {'A':22,'B':14,'C':14,'D':14,'E':14,'F':12,'G':12,'H':12,'I':12,'J':12})

title_row(ws_comp, 1, '  DIXON TECHNOLOGIES — PEER COMPARABLES  |  EMS / Electronics Manufacturing India  (FY25A, ₹ Cr)', cols=10)

ws_comp.merge_cells('A3:J3')
src_note = ws_comp.cell(row=3, column=1,
    value="Sources: BSE/NSE filings, Screener.in, Trendlyne  |  Market cap as of June 2026 (approx.)  |  "
          "All figures in ₹ Crore unless stated")
src_note.font = Font(name='Calibri', italic=True, color="595959", size=9)
src_note.alignment = al('left', indent=1)

# Column headers
comp_hdrs = ['Company', 'Revenue', 'EBITDA', 'EBITDA %', 'PAT', 'Mkt Cap', 'Net Debt', 'EV', 'EV/EBITDA', 'P/E']
col_hdr(ws_comp, 5, comp_hdrs, start_col=1, bg=DARK_BLUE)

peers = [
    # name,  rev,   ebitda, ebitda_pct, pat,  mcap,  net_debt, ev,    ev_ebitda, pe
    ("Dixon Technologies",    38860, 1943, 0.050, 1215, 75000, 200,   75200, 38.7, 61.7),
    ("Amber Enterprises",      5200,  442, 0.085,  148, 20000, 600,   20600, 46.6, 135.0),
    ("Kaynes Technology",      1800,  270, 0.150,  175, 14000, 150,   14150, 52.4, 80.0),
    ("Syrma SGS Technology",   2800,  196, 0.070,  100,  6500, 350,    6850, 34.9, 65.0),
]

for ri, (name, rev, ebitda, epct, pat, mcap, nd, ev, eve, pe) in enumerate(peers, start=6):
    is_dixon = ri == 6
    bg_c = LIGHT_YELLOW if is_dixon else (XLIGHT_BLUE if ri % 2 == 0 else None)
    fw = True if is_dixon else False

    data = [name, rev, ebitda, epct, pat, mcap, nd, ev, eve, pe]
    fmts = [None, INR, INR, PCT, INR, INR, INR, INR, MULT, MULT]

    for ci, (val, fmt) in enumerate(zip(data, fmts), start=1):
        c = ws_comp.cell(row=ri, column=ci, value=val)
        c.font = Font(name='Calibri', bold=fw, size=10, color="000000")
        if fmt:
            if fmt == PCT:
                c.alignment = al('center')
            elif fmt in [MULT]:
                c.alignment = al('center')
            else:
                c.alignment = al('right')
            c.number_format = fmt
        else:
            c.alignment = al('left', indent=1)
        if bg_c:
            c.fill = fill(bg_c)

# Separator
ws_comp.row_dimensions[10].height = 6

# Stats rows
def comps_stat(ws, row, label, formula_col_range, fmt, bold=False, bg=None):
    c = ws.cell(row=row, column=1, value=label)
    c.font = Font(name='Calibri', bold=bold, size=10)
    c.alignment = al('left', indent=1)
    if bg:
        c.fill = fill(bg)
    for ci, col_letter in enumerate(formula_col_range, start=2):
        row_range = f"{col_letter}7:{col_letter}9"  # peers only (rows 7-9, excluding Dixon row 6)
        val = f'=MEDIAN({row_range})'
        if 'MAX' in label:
            val = f'=MAX({row_range})'
        elif 'MIN' in label:
            val = f'=MIN({row_range})'
        cell = ws.cell(row=row, column=ci, value=val)
        cell.number_format = fmt[ci-2] if isinstance(fmt, list) else fmt
        cell.alignment = al('center')
        cell.font = Font(name='Calibri', bold=bold, size=10)
        if bg:
            cell.fill = fill(bg)

stat_fmts = [INR, INR, PCT, INR, INR, INR, INR, MULT, MULT]
col_letters = ['B','C','D','E','F','G','H','I','J']

for ri, (label, bold, bg) in enumerate([
    ("Peer Median  (ex-Dixon)", True, LIGHT_BLUE),
    ("Peer Mean   (ex-Dixon)",  False, XLIGHT_BLUE),
], start=11):
    ws_comp.cell(row=ri, column=1, value=label).font = Font(name='Calibri', bold=bold, size=10)
    ws_comp.cell(row=ri, column=1).alignment = al('left', indent=1)
    if bg:
        ws_comp.cell(row=ri, column=1).fill = fill(bg)
    for ci, (cl, fmt) in enumerate(zip(col_letters, stat_fmts), start=2):
        row_range = f"{cl}7:{cl}9"
        func = "MEDIAN" if ri == 11 else "AVERAGE"
        cell = ws_comp.cell(row=ri, column=ci, value=f'={func}({row_range})')
        cell.number_format = fmt
        cell.font = Font(name='Calibri', bold=bold, size=10)
        cell.alignment = al('right') if fmt != PCT and fmt != MULT else al('center')
        if bg:
            cell.fill = fill(bg)

# Dixon vs Peer comparison notes
ws_comp.row_dimensions[14].height = 6
sub_hdr(ws_comp, 15, '  POSITIONING ANALYSIS', bg=MED_BLUE, cols=10)

notes = [
    "► Dixon's EBITDA margin (5.0%) is below EMS peers due to high mobile (low-margin) revenue mix — Kaynes (15%) is PCBA-focused (higher value-add)",
    "► Dixon trades at 38.7x EV/EBITDA vs. peer median ~46-52x, suggesting relative value or market pricing in lower margin profile",
    "► Kaynes & Amber have better margins but much smaller scale — Dixon is the only large-cap pure EMS play in India",
    "► China+1, PLI scheme benefits, and new category wins (IT hardware, wearables) are key re-rating catalysts for Dixon",
    "► Key risk: margin compression if mobile mix rises further; PLI scheme sunset risk post FY27",
]
for i, note in enumerate(notes, start=16):
    ws_comp.merge_cells(start_row=i, start_column=1, end_row=i, end_column=10)
    c = ws_comp.cell(row=i, column=1, value=note)
    c.font = Font(name='Calibri', size=10)
    c.alignment = al('left', indent=1)
    if i % 2 == 0:
        c.fill = fill(XLIGHT_BLUE)

# Update Cover nav to reflect 8 tabs (no WACC tab)
# Already done in Cover build above

# ─────────────────────────────────────────────────────────────────────────────
# GLOBAL FREEZE & SAVE
# ─────────────────────────────────────────────────────────────────────────────
for ws in [ws_is, ws_bs, ws_cf]:
    ws.freeze_panes = 'B5'

out_path = "/sessions/clever-zen-darwin/mnt/outputs/Dixon_3Statement_DCF_Model_v2.0.xlsx"
wb.save(out_path)
print(f"✅  Saved: {out_path}")
print(f"   Tabs: {[s.title for s in wb.worksheets]}")
